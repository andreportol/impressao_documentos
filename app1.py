from PyPDF3 import PdfFileReader
from docx2pdf import convert
from openpyxl import Workbook
import win32print
import win32api
import os



class App1:

    # criei manualmente a pasta "pdf" dentro da pasta "teste"

    def converter_arquivos(self):
        self.caminho1 = r"C:\Users\andre.porto\Desktop\teste"
        self.caminho2 = r"C:\Users\andre.porto\Desktop\teste\pdf"
        # conversão em massa dos arquivos "docx" para pasta "pdf"
        convert(self.caminho1,
                self.caminho2)

    def criar_planilha(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "ANDRÉ PORTO"
        self.ws.sheet_properties.tabColor = "1072BA"
        self.ws['A1'] = 'ARQUIVOS'
        self.ws['B1'] = 'N. DE CÓPIAS'
        self.ws['C1'] = 'N. DE PÁGINAS'

    def preencher_planilha_controle(self):
        # path dos documentos em"pdf" a serem carregados
        # self.caminho = r"C:\Users\Thay\Desktop\teste\pdf"
        # criando uma lista com os arquivos obtidos dentro da pasta 'pdf'
        lista_de_arquivos = os.listdir(self.caminho2)
        lista_de_arquivos.sort() # coloca a lista em ordem alfabetica
        c = 1  # variável criada para ler a linha do arquivo em excel
        for arquivos in lista_de_arquivos:
            pdf = PdfFileReader(
                open(self.caminho2+'\\' + arquivos, 'rb'))
            numero = int(pdf.getNumPages())
            c += 1
            self.ws.cell(row=c, column=1, value=arquivos)
            self.ws.cell(row=c, column=2, value=1)
            self.ws.cell(row=c, column=3, value=numero)
            # Save the file
            self.wb.save(self.caminho1 + r'\controle.xlsx')

    
if __name__ == '__main__':
    aplicativo = App1()
    aplicativo.converter_arquivos()
    aplicativo.criar_planilha()
    aplicativo.preencher_planilha_controle()
    print(' FIM -> PROGRAMA FINALIZADO....PLANILHA CARREGADA!!')
    # teste commit